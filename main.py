import pandas as pd
import os
import webbrowser
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import math
import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk  # For custom icons

  
# Mapping dictionary for product names and packaging info
PRODUCT_INFO = {
    "Finesse 5L": {"mapping": "FINESSE 5L JMAL", "unit": "Ct", "pieces_per_unit": 3},
    "Finesse 3L": {"mapping": "FINESSE 3L JMAL", "unit": "Ct", "pieces_per_unit": 4},
    "Finesse 2L": {"mapping": "FINESSE 2L JMAL", "unit": "Ct", "pieces_per_unit": 6},
    "Finesse 1L": {"mapping": "FINESSE 1L JMAL", "unit": "Ct", "pieces_per_unit": 12},
    "Class 5L": {"mapping": "CLASS LINGE BLEU 5L JMAL", "unit": "Ct", "pieces_per_unit": 3},
    "Class 3L": {"mapping": "CLASS LINGE 3L JMAL", "unit": "Ct", "pieces_per_unit": 4},
    "Class 2L": {"mapping": "CLASS LINGE 2L JMAL", "unit": "Ct", "pieces_per_unit": 6},
    "Degraissant super mag": {"mapping": "CLEAN DEGRAISSANT", "unit": "Lot", "pieces_per_unit": 6},
    "Line 2L": {"mapping": "LINE 2L PECHE JMAL", "unit": "Ct", "pieces_per_unit": 6},
    "Fawah 5L": {"mapping": "FAWAH 5L JMAL", "unit": "Ct", "pieces_per_unit": 3},
    "Fawah 2L": {"mapping": "FAWAH 2L JMAL", "unit": "Ct", "pieces_per_unit": 6},
    "Fawah 90cL": {"mapping": "FAWAH 1L JMAL", "unit": "Ct", "pieces_per_unit": 12},
    "Ghassel 5L": {"mapping": "GHASSEL 5L JMAL", "unit": "Ct", "pieces_per_unit": 3},
    "Ghassel 2L": {"mapping": "GHASSEL 2L JMAL", "unit": "Ct", "pieces_per_unit": 6},
    "Ghassel wc": {"mapping": "GHASSEL WC 1L JMAL", "unit": "Ct", "pieces_per_unit": 12},
    "Alys 5L": {"mapping": "ALYS 5L POMME JMAL", "unit": "Ct", "pieces_per_unit": 3},
    "Alys 2L": {"mapping": "ALYS 2L POMME JMAL", "unit": "Ct", "pieces_per_unit": 6},
    "Alys 1L": {"mapping": "ALYS 1L POMME JMAL", "unit": "Ct", "pieces_per_unit": 12},
    "Javel 5L": {"mapping": "JAVEL JMAL 5L", "unit": "Ct", "pieces_per_unit": 3},
    "Javel 2L": {"mapping": "JAVEL JMAL 2L", "unit": "Ct", "pieces_per_unit": 6},
    "Javel 90cL": {"mapping": "JAVEL JMAL 1L", "unit": "Ct", "pieces_per_unit": 12},
    "Bassatine 5L": {"mapping": "BASSATINE 5L BLEU JMAL", "unit": "Ct", "pieces_per_unit": 3},
    "Bassatine 1L": {"mapping": "BASSATINE 1L BLEU JMAL", "unit": "Ct", "pieces_per_unit": 12},
    "Vinaigre": {"mapping": "VINAIGRE BLANC", "unit": "Lot", "pieces_per_unit": 6},
    "Balais toscana": {"mapping": "BALAIS TOSCANA 12PC", "unit": "Ct", "pieces_per_unit": 12},
    "Balais flouka": {"mapping": "BALAI FLOUKA", "unit": "Ps", "pieces_per_unit": 1},
    "SAC CUISSANT": {"mapping": "SAC CUISSON SPONGEX", "unit": "Ps", "pieces_per_unit": 1},
    "Aluminium Novalu": {"mapping": "NOVALU 500GR", "unit": "Ct", "pieces_per_unit": 6},
    "Etirable Ref 300": {"mapping": "ETIRABLE DIAMANT", "unit": "Ct", "pieces_per_unit": 12},
    "Aluminium 8m": {"mapping": "ALLUMINIUM 8M", "unit": "Lot", "pieces_per_unit": 10},
    "Alumium ref 8": {"mapping": "ALLUMINIUM 5M", "unit": "Lot", "pieces_per_unit": 10},
    "Etirable 8m": {"mapping": "ETIRABLE 8M", "unit": "Lot", "pieces_per_unit": 10},
    "Diptox GM": {"mapping": "DIPTOX GM", "unit": "Ct", "pieces_per_unit": 12},
    "Diptox PM": {"mapping": "DIPTOX PM", "unit": "Ct", "pieces_per_unit": 12},
    "Choc Combat GM": {"mapping": "CHOC COMBAT GM", "unit": "Ct", "pieces_per_unit": 10},
    "Raclette dumax": {"mapping": "RACLETTE DUMAX 45 GOLDEN", "unit": "Ps", "pieces_per_unit": 1},
    "SAC GEANT SUPERT": {"mapping": "SAC GEANT SUPER", "unit": "Lot", "pieces_per_unit": 1},
    "Serpillère GM": {"mapping": "SERPIERRE BLANC GM", "unit": "Lot", "pieces_per_unit": 1},
    "Sonit": {"mapping": "SERPIERRE SONIT", "unit": "Lot", "pieces_per_unit": 1},
    "Jax carré super": {"mapping": "JAX CARRE DE 10", "unit": "Lot", "pieces_per_unit": 1},
    "Jax Fer JEX": {"mapping": "JAX FER SUPER", "unit": "Lot", "pieces_per_unit": 1},
    "Lavette microfibre": {"mapping": "LAVETTE MICROFIBRE BINGO", "unit": "Ps", "pieces_per_unit": 1},
    "Eponge double face": {"mapping": "EPONGE DOUBLE FACE PM", "unit": "Lot", "pieces_per_unit": 1},
    "Plaque inox": {"mapping": "PLAQUE INOX DE 12", "unit": "Lot", "pieces_per_unit": 12},
    "Jax inox de 3": {"mapping": "JAX INOX DE 3", "unit": "Ps", "pieces_per_unit": 1},
    "Dexel 10 Kg": {"mapping": "DEXEL 10KG", "unit": "Sac", "pieces_per_unit": 1},
    "Pince linge cobra": {"mapping": "PINCE COBRA", "unit": "Ps", "pieces_per_unit": 1},
    "Cachemir": {"mapping": "CACHMIR BLANC", "unit": "Lot", "pieces_per_unit": 1},
    "Sunsilk": {"mapping": "SUNSILK 350ML", "unit": "Ct", "pieces_per_unit": 12},
    "MANCHE METALLIQ 1.2": {"mapping": "MANCHE METALIQUE 1.2", "unit": "Lot", "pieces_per_unit": 12},
    "Savon DUMAX 400ML": {"mapping": "DUMAX 400ML", "unit": "Ct", "pieces_per_unit": 24},
    "Savon DUMAX 1L": {"mapping": "SAVON LIQUIDE DUMAX 1L", "unit": "Ct", "pieces_per_unit": 12},
    "MANCHE METALLIQ 1.4": {"mapping": "MANCHE METALIQUE 1.4", "unit": "Lot", "pieces_per_unit": 12},
    "CHOC WC": {"mapping": "CHOC WC", "unit": "Ct", "pieces_per_unit": 12},
    "PAPIER CUISSANT": {"mapping": "PAPIER CUISSANT", "unit": "Ps", "pieces_per_unit": 1},
    "MANCHE BOIS 1.4": {"mapping": "MANCHE BOIS 1.4", "unit": "Lot", "pieces_per_unit": 12},
    "Pince papillon": {"mapping": "PINCE PAPILLON", "unit": "Ps", "pieces_per_unit": 1},
    "Sachet congelation": {"mapping": "SAC CONGELATION 1L", "unit": "Ct", "pieces_per_unit": 50},
    "Lexus vitre": {"mapping": "LUXIS VITRE", "unit": "Ct", "pieces_per_unit": 10},
    "Gant menage zony": {"mapping": "GANT MENAGE L", "unit": "Lot", "pieces_per_unit": 12},
    "Charbon": {"mapping": "CHARBON", "unit": "Ct", "pieces_per_unit": 60},
    "Air'fresh plein air": {"mapping": "AIR FRESH PLEIN AIR", "unit": "Ct", "pieces_per_unit": 12},
    "Manche BOIS 1.2": {"mapping": "MANCHE BOIS 1.2", "unit": "Lot", "pieces_per_unit": 12},
    "spontex": {"mapping": "SPONTEX", "unit": "Lot", "pieces_per_unit": 1},
    "Sac poubelle 50/80 n": {"mapping": "SAC 50/80 NORMAL", "unit": "Sac", "pieces_per_unit": 1},
    "Tej 375gr": {"mapping": "TEJ 375 GR VERT", "unit": "Ct", "pieces_per_unit": 24},
    "Zitouna 100gr": {"mapping": "ZITOUNA 100GR", "unit": "Ct", "pieces_per_unit": 50},
    "Serpierre raclette": {"mapping": "SERPILLIERE RACLETTE", "unit": "Ps", "pieces_per_unit": 1},
    "Serpierre microfibre": {"mapping": "SERPILLIERE MICROFIBRE BINGO", "unit": "Ps", "pieces_per_unit": 1},
    "Zitouna 400gr": {"mapping": "ZITOUNA 400GR VERT", "unit": "Ct", "pieces_per_unit": 24},
    "Tej 125gr": {"mapping": "TEJ 100 GR VERT", "unit": "Ct", "pieces_per_unit": 72},
    "50*80 super": {"mapping": "SAC 50/80 SUPER", "unit": "Sac", "pieces_per_unit": 1},
    "Poubelle geant": {"mapping": "SAC GEANT NORMAL", "unit": "Sac", "pieces_per_unit": 1},
    "JAX RENFORCE": {"mapping": "JAX RENFORCE", "unit": "Lot", "pieces_per_unit": 1},
    "CURDENT": {"mapping": "CURDENT", "unit": "Lot", "pieces_per_unit": 1},
    "CORDE A LINGE": {"mapping": "CORDE A LINGE", "unit": "Lot", "pieces_per_unit": 10},
    "PASTILLE HACKER": {"mapping": "PASTILLE HAKER", "unit": "Lot", "pieces_per_unit": 10}
}

# def format_quantity(total, product_info):
#     if pd.isna(total) or total == 0:
#         return ""
    
#     unit = product_info["unit"]
#     pieces_per_unit = product_info["pieces_per_unit"]
    
#     if unit in ["Ps", "Sac", "Lot"] and pieces_per_unit == 1:
#         return f"{int(total)} {unit}"
    
#     full_units = total // pieces_per_unit
#     remaining_pieces = total % pieces_per_unit
    
#     if full_units > 0 and remaining_pieces > 0:
#         return f"{int(full_units)} {unit} + {int(remaining_pieces)} Ps"
#     elif full_units > 0:
#         return f"{int(full_units)} {unit}"
#     else:
#         return f"{int(remaining_pieces)} Ps"

def format_quantity(total, product_info):
    if pd.isna(total) or total == 0:
        return ""
    
    unit = product_info["unit"]
    pieces_per_unit = product_info["pieces_per_unit"]
    
    # Handle negative values
    is_negative = total < 0
    absolute_total = abs(total)
    
    if unit in ["Ps", "Sac", "Lot"] and pieces_per_unit == 1:
        return f"{'-' if is_negative else ''}{int(absolute_total)} {unit}"
    
    full_units = absolute_total // pieces_per_unit
    remaining_pieces = absolute_total % pieces_per_unit
    
    parts = []
    if full_units > 0:
        parts.append(f"{full_units} {unit}")
    if remaining_pieces > 0:
        parts.append(f"{remaining_pieces} Ps")
    
    formatted = " + ".join(parts) if parts else "0"
    
    # Add negative sign if needed
    if is_negative:
        formatted = f"-{formatted}"
    
    return formatted

def compare_excel(file1_path, file2_path, output_path):
    try:
        # Read files
        df1 = pd.read_excel(file1_path, engine='xlrd')
        df2 = pd.read_excel(file2_path, engine='xlrd')
    
        # Validate required columns exist
        required_cols_file1 = ["Designation", "Qté"]
        required_cols_file2 = ["designation", "Qte Vente"]

        if not all(col in df1.columns for col in required_cols_file1):
            raise ValueError("File 1 is missing required columns (Designation, Qté)")
        if not all(col in df2.columns for col in required_cols_file2):
            raise ValueError("File 2 is missing required columns (designation, Qte Vente)")
        
        # First part: Products in file2 (with possible matches in file1)
        merged_df = df2.merge(
            df1,
            left_on="designation",
            right_on="Designation",
            how="left"
        ).fillna(0)  # Replace NaN with 0

        # Create main result dataframe
        main_result = pd.DataFrame()
        main_result["name of the product"] = merged_df["designation"]
        main_result["quantity"] = merged_df["Qte Vente"]
        main_result["ajout"] = merged_df["Qté"].astype(int)
        main_result["difference"] = main_result["quantity"] - main_result["ajout"]

        # Second part: Products only in file1
        only_in_file1 = df1[~df1["Designation"].isin(df2["designation"])].copy()
        only_in_file1_result = pd.DataFrame()
        only_in_file1_result["name of the product"] = only_in_file1["Designation"]
        only_in_file1_result["quantity"] = 0
        only_in_file1_result["ajout"] = only_in_file1["Qté"].astype(int)
        only_in_file1_result["difference"] = only_in_file1_result["quantity"] - only_in_file1_result["ajout"]

        # Third part: 33333 dataframe with specified products
        df_33333 = pd.DataFrame({
            "Désignation": list(PRODUCT_INFO.keys()),
            "Total": None  # Initialize as empty
        })
        
        # Create a dictionary for quick lookup from main_result
        main_result_dict = dict(zip(
            main_result["name of the product"].str.strip().str.upper(),
            main_result["difference"]
        ))
        
        # Fill in the totals from main comparison and format them
        matched_products = []
        unmatched_products = []
        
        for idx, row in df_33333.iterrows():
            product_name = row["Désignation"]
            product_info = PRODUCT_INFO[product_name]
            mapped_name = product_info["mapping"].strip().upper()
            
            if mapped_name in main_result_dict:
                total = main_result_dict[mapped_name]

                if total < 0 or total == 0:
                    continue

                formatted_total = format_quantity(total, product_info)
                df_33333.at[idx, "Total"] = formatted_total
                matched_products.append(product_name)
            else:
                # Try again with original mapping (without stripping) for backward compatibility
                original_mapped_name = product_info["mapping"].upper()
                if original_mapped_name in main_result_dict:
                    total = main_result_dict[original_mapped_name]

                    if total < 0 or total == 0:
                        continue

                    formatted_total = format_quantity(total, product_info)
                    df_33333.at[idx, "Total"] = formatted_total
                    matched_products.append(product_name)
                else:
                    unmatched_products.append(product_name)
        
        # Create separate dataframes for matched and unmatched products
        matched_df = df_33333[df_33333["Désignation"].isin(matched_products)].copy()
        unmatched_df = df_33333[df_33333["Désignation"].isin(unmatched_products)].copy()

        # Save to Excel with multiple sheets
        with pd.ExcelWriter(output_path) as writer:
            main_result.to_excel(writer, sheet_name="Main Comparison", index=False)
            only_in_file1_result.to_excel(writer, sheet_name="Only in Ajout", index=False)
            matched_df.to_excel(writer, sheet_name="33333 Matched", index=False)
            unmatched_df.to_excel(writer, sheet_name="33333 Unmatched", index=False)

        # Apply formatting
        wb = load_workbook(output_path)
        
        # Format main comparison sheet
        ws_main = wb["Main Comparison"]
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for row in range(2, ws_main.max_row + 1):
            if ws_main.cell(row=row, column=4).value < 0:
                ws_main.cell(row=row, column=4).fill = red_fill

        # Format only in ajout sheet (blue)
        ws_ajout = wb["Only in Ajout"]
        blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        for row in range(2, ws_ajout.max_row + 1):
            ws_ajout.cell(row=row, column=4).fill = blue_fill

        # Format 33333 Matched sheet (green for matched products)
        ws_33333_matched = wb["33333 Matched"]
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        for row in range(2, ws_33333_matched.max_row + 1):
            if ws_33333_matched.cell(row=row, column=2).value is not None:
                ws_33333_matched.cell(row=row, column=2).fill = green_fill

        # Format 33333 Unmatched sheet (yellow for unmatched products)
        ws_33333_unmatched = wb["33333 Unmatched"]
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        for row in range(2, ws_33333_unmatched.max_row + 1):
            ws_33333_unmatched.cell(row=row, column=1).fill = yellow_fill
            ws_33333_unmatched.cell(row=row, column=2).fill = yellow_fill

        # Save the workbook with formatting
        wb.save(output_path)
        print(f"Output saved to: {output_path} with four sheets")

    except Exception as e:
        raise ValueError(f"Error processing files: {str(e)}")
    
def create_gui():
    root = tk.Tk()
    root.title("Excel Comparator Pro")
    
    # Set window icon (replace with your icon file)
    try:
        root.iconbitmap('icon.ico')  # Place your .ico file in same folder
    except:
        pass  # Skip if icon file not found
    
    root.geometry("600x400")
    
    # Custom style
    root.configure(bg='#f0f0f0')
    style = {'font': ('Arial', 10), 'bg': '#f0f0f0', 'fg': '#333333'}
    highlight_style = {'font': ('Arial', 10, 'bold'), 'bg': '#f0f0f0', 'fg': '#0066cc'}
    
    # Header frame
    header_frame = tk.Frame(root, bg='#4b8bbe')
    header_frame.pack(fill=tk.X)
    
    # Add your photo (replace with your image file)
    try:
        photo = Image.open("houssem_photo.png").resize((50,50))
        photo_img = ImageTk.PhotoImage(photo)
        photo_label = tk.Label(header_frame, image=photo_img, bg='#4b8bbe')
        photo_label.image = photo_img  # Keep reference
        photo_label.pack(side=tk.LEFT, padx=10, pady=5)
    except:
        pass  # Skip if image not found
    
    # Title
    title = tk.Label(header_frame, 
                   text="الروكوبومون", 
                   font=('Arial', 20, 'bold'), 
                   bg='#4b8bbe', fg='white')
    title.pack(side=tk.LEFT, padx=100)
    
    # Instructions frame
    instructions_frame = tk.Frame(root, bg='#f0f0f0')
    instructions_frame.pack(pady=20)
    
    # Step-by-step instructions
    steps = [
        "1. First select the AJOUT file",
        "2. Then select the VENTE file",
        "3. Finally choose where to save your comparison result"
    ]
    
    for step in steps:
        tk.Label(instructions_frame, 
                text=step, 
                **style, 
                justify=tk.LEFT).pack(anchor=tk.W, pady=5)
    
    # Main button
    button_frame = tk.Frame(root, bg='#f0f0f0')
    button_frame.pack(pady=20)
    
    run_btn = tk.Button(button_frame, 
                       text="START COMPARISON", 
                       command=run_comparison, 
                       height=2, 
                       width=25,
                       bg='#4CAF50',  # Green color
                       fg='white',
                       font=('Arial', 12, 'bold'),
                       relief=tk.RAISED,
                       bd=3)
    run_btn.pack()
    
    # Footer with credits
    footer_frame = tk.Frame(root, bg='#f0f0f0')
    footer_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=10)
    
    def open_channel():
        webbrowser.open("https://www.youtube.com/watch?v=3UE5jBg6JRg")
    
    tk.Label(footer_frame, 
             text="Made by ", 
             **style).pack(side=tk.LEFT, padx=10)
    
    name_label = tk.Label(footer_frame, 
                         text="Jallouli Houssem-Eddin", 
                         cursor="hand2", 
                         **highlight_style)
    name_label.pack(side=tk.LEFT)
    name_label.bind("<Button-1>", lambda e: open_channel())
    
    version_label = tk.Label(footer_frame, 
                            text="v1.0", 
                            **style)
    version_label.pack(side=tk.RIGHT, padx=10)
    
    root.mainloop()

def run_comparison():
    # [Keep your existing run_comparison function...]
    file1 = filedialog.askopenfilename(
        title="Select File 1 (Ajout)",
        filetypes=[("Excel files", "*.xls *.xlsx")]
    )
    if not file1:
        return
        
    file2 = filedialog.askopenfilename(
        title="Select File 2 (Vente)",
        filetypes=[("Excel files", "*.xls *.xlsx")]
    )
    if not file2:
        return
        
    output = filedialog.asksaveasfilename(
        title="Save Output As",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not output:
        return

    try:
        compare_excel(file1, file2, output)
        messagebox.showinfo("Success", f"File processed successfully!\nSaved to: {output}")
    except Exception as e:
        messagebox.showerror("Error", f"Processing failed:\n{str(e)}")

if __name__ == "__main__":
    create_gui()
